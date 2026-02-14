"""
Quote Views for MCD-Agencia.

This module provides ViewSets for quote operations:
    - Quote request submission (public)
    - Quote management (admin)
    - Quote acceptance (customer)
"""

import logging
import threading
from decimal import Decimal
from django.db import models, transaction
from django.db.models import Count, Sum, Q
from django.http import FileResponse, Http404, HttpResponse
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.filters import SearchFilter
from rest_framework.response import Response
from rest_framework.views import APIView
from django_filters.rest_framework import DjangoFilterBackend

logger = logging.getLogger(__name__)

from apps.audit.models import AuditLog
from apps.core.pagination import StandardPagination
from apps.core.views import is_internal_user as _is_internal_user
from apps.notifications.models import Notification
from .models import QuoteRequest, Quote, QuoteLine, QuoteAttachment, QuoteChangeRequest, QuoteResponse
from .tasks import send_quote_email_sync, notify_admin_new_quote_request, send_quote_accepted_notification, send_order_confirmation_email
from .serializers import (
    QuoteRequestSerializer,
    QuoteRequestCreateSerializer,
    QuoteRequestAdminSerializer,
    QuoteSerializer,
    QuoteAdminSerializer,
    QuoteCreateSerializer,
    QuoteLineSerializer,
    QuoteSendSerializer,
    QuoteAcceptSerializer,
    SalesRepDashboardSerializer,
    QuoteChangeRequestCreateSerializer,
    QuoteChangeRequestSerializer,
    QuoteChangeRequestReviewSerializer,
)


class QuoteRequestPublicView(APIView):
    """
    Public endpoint for quote request submission.

    POST /api/v1/quotes/request/
    """

    permission_classes = [permissions.AllowAny]

    def post(self, request):
        """Submit a new quote request."""
        serializer = QuoteRequestCreateSerializer(
            data=request.data,
            context={'request': request}
        )
        serializer.is_valid(raise_exception=True)
        quote_request = serializer.save()

        # Log creation
        AuditLog.log(
            entity=quote_request,
            action=AuditLog.ACTION_CREATED,
            after_state=QuoteRequestSerializer(quote_request).data,
            request=request,
            metadata={'source': 'public_form'}
        )

        # Notify sales team about new quote request
        try:
            notify_admin_new_quote_request.delay(str(quote_request.id))
        except Exception:
            pass  # Don't fail the request if notification fails

        # In-app notification: targeted based on urgency
        try:
            if quote_request.urgency == QuoteRequest.URGENCY_HIGH:
                # Urgent → notify ALL staff (admin + sales)
                Notification.notify_staff(
                    notification_type=Notification.TYPE_QUOTE_REQUEST,
                    title=f'🔴 Solicitud urgente #{quote_request.request_number}',
                    message=f'{quote_request.customer_name} — {quote_request.description[:100]}',
                    entity_type='QuoteRequest',
                    entity_id=quote_request.id,
                    action_url=f'/dashboard/solicitudes/{quote_request.id}',
                )
            else:
                # Normal/Medium → notify assigned seller + admins
                Notification.notify_assigned_seller_and_admins(
                    assigned_to=quote_request.assigned_to,
                    notification_type=Notification.TYPE_QUOTE_REQUEST,
                    title=f'Nueva solicitud #{quote_request.request_number}',
                    message=f'{quote_request.customer_name} — {quote_request.description[:100]}',
                    entity_type='QuoteRequest',
                    entity_id=quote_request.id,
                    action_url=f'/dashboard/solicitudes/{quote_request.id}',
                )
        except Exception:
            pass

        return Response(
            {
                'message': _('Your quote request has been submitted. We will contact you soon.'),
                'request_number': quote_request.request_number
            },
            status=status.HTTP_201_CREATED
        )


class QuoteRequestViewSet(viewsets.ModelViewSet):
    """
    ViewSet for quote request management.

    Customer endpoints:
        GET /api/v1/quotes/requests/ - List user's requests
        GET /api/v1/quotes/requests/{id}/ - Request details

    Admin endpoints:
        GET /api/v1/admin/quote-requests/ - List all requests
        PUT /api/v1/admin/quote-requests/{id}/ - Update request
        POST /api/v1/admin/quote-requests/{id}/assign/ - Assign to sales
    """

    permission_classes = [permissions.IsAuthenticated]
    pagination_class = StandardPagination
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['status', 'assigned_to', 'urgency']
    search_fields = ['request_number', 'customer_name', 'customer_email', 'customer_company', 'catalog_item_name']

    def get_queryset(self):
        """Return requests based on user role and visibility rules.

        Visibility rules:
            - Admin: sees ALL requests.
            - Sales: sees requests assigned to them + high urgency requests.
            - Regular customer: only their own requests (by email).

        List-view default: hide already-quoted / accepted / rejected requests
        (they live in the Cotizaciones section). The filter is skipped when:
            - an explicit ?status= query-param is provided, OR
            - the action is 'retrieve' (detail view).
        """
        user = self.request.user
        base_qs = QuoteRequest.objects.select_related(
            'catalog_item', 'assigned_to'
        ).prefetch_related('attachments')

        if _is_internal_user(user):
            # Sales reps: only assigned to them + high urgency
            if hasattr(user, 'role') and user.role and user.role.name == 'sales':
                base_qs = base_qs.filter(
                    models.Q(assigned_to=user) |
                    models.Q(urgency=QuoteRequest.URGENCY_HIGH)
                )
            # else: admins see everything (no extra filter)

            # On list action, hide already-processed requests unless the
            # user explicitly filters by status.
            if self.action == 'list' and 'status' not in self.request.query_params:
                base_qs = base_qs.exclude(
                    status__in=[
                        QuoteRequest.STATUS_QUOTED,
                        QuoteRequest.STATUS_ACCEPTED,
                        QuoteRequest.STATUS_REJECTED,
                    ]
                )

            return base_qs

        # Regular customer sees their own requests
        return base_qs.filter(
            customer_email=user.email,
            is_deleted=False
        )

    def get_serializer_class(self):
        """Return appropriate serializer."""
        if _is_internal_user(self.request.user):
            return QuoteRequestAdminSerializer
        return QuoteRequestSerializer

    @action(detail=False, methods=['get'])
    def dashboard(self, request):
        """Get sales rep dashboard data."""
        if not _is_internal_user(request.user):
            return Response(status=status.HTTP_403_FORBIDDEN)

        user = request.user

        # Get requests based on role visibility rules
        if hasattr(user, 'role') and user.role and user.role.name == 'sales':
            # Sales rep: stats scoped to assigned requests + their quotes
            requests_qs = QuoteRequest.objects.filter(
                models.Q(assigned_to=user) | models.Q(urgency=QuoteRequest.URGENCY_HIGH)
            )
            quotes_qs = Quote.objects.filter(created_by=user)
        else:
            requests_qs = QuoteRequest.objects.all()
            quotes_qs = Quote.objects.all()

        # Calculate stats
        pending_requests = requests_qs.filter(
            status__in=[QuoteRequest.STATUS_PENDING, QuoteRequest.STATUS_ASSIGNED, QuoteRequest.STATUS_IN_REVIEW]
        ).count()

        quotes_without_response = quotes_qs.filter(
            status__in=[Quote.STATUS_SENT, Quote.STATUS_VIEWED]
        ).count()

        total_quotes = quotes_qs.exclude(status=Quote.STATUS_DRAFT).count()
        accepted_quotes = quotes_qs.filter(status=Quote.STATUS_ACCEPTED).count()

        conversion_rate = Decimal('0.00')
        if total_quotes > 0:
            conversion_rate = Decimal(str(round((accepted_quotes / total_quotes) * 100, 2)))

        total_quoted = quotes_qs.exclude(
            status=Quote.STATUS_DRAFT
        ).aggregate(total=Sum('total'))['total'] or Decimal('0.00')

        total_approved = quotes_qs.filter(
            status=Quote.STATUS_ACCEPTED
        ).aggregate(total=Sum('total'))['total'] or Decimal('0.00')

        # Get urgent requests
        urgent_requests = requests_qs.filter(
            urgency=QuoteRequest.URGENCY_HIGH,
            status__in=[QuoteRequest.STATUS_PENDING, QuoteRequest.STATUS_ASSIGNED, QuoteRequest.STATUS_IN_REVIEW]
        ).order_by('-created_at')[:5]

        # Get recent activity from audit log
        activity_qs = AuditLog.objects.filter(
            entity_type__in=['Quote', 'QuoteRequest', 'Order'],
        ).order_by('-timestamp')[:15]

        recent_activity = []
        for log in activity_qs:
            description = log.entity_repr or f"{log.entity_type}"
            action_label = log.get_action_display()
            recent_activity.append({
                'id': str(log.id),
                'action': log.action,
                'action_display': action_label,
                'entity_type': log.entity_type,
                'entity_id': log.entity_id,
                'description': f"{description} — {action_label}",
                'actor': log.actor_email or 'Sistema',
                'timestamp': log.timestamp.isoformat(),
                'metadata': log.metadata,
            })

        data = {
            'pending_requests': pending_requests,
            'quotes_without_response': quotes_without_response,
            'conversion_rate': conversion_rate,
            'total_quoted': total_quoted,
            'total_approved': total_approved,
            'urgent_requests': QuoteRequestSerializer(urgent_requests, many=True).data,
            'recent_activity': recent_activity,
        }

        return Response(data)

    @action(detail=True, methods=['post'])
    def assign(self, request, pk=None):
        """Assign quote request to sales rep (admin)."""
        if not _is_internal_user(request.user):
            return Response(status=status.HTTP_403_FORBIDDEN)

        quote_request = self.get_object()
        assigned_to_id = request.data.get('assigned_to_id')

        if not assigned_to_id:
            return Response(
                {'error': _('assigned_to_id is required.')},
                status=status.HTTP_400_BAD_REQUEST
            )

        from django.contrib.auth import get_user_model
        User = get_user_model()

        try:
            assigned_to = User.objects.get(id=assigned_to_id, is_active=True)
        except User.DoesNotExist:
            return Response(
                {'error': _('User not found.')},
                status=status.HTTP_404_NOT_FOUND
            )

        quote_request.assigned_to = assigned_to
        quote_request.assigned_at = timezone.now()
        # Update status to assigned if currently pending
        if quote_request.status in [QuoteRequest.STATUS_PENDING, QuoteRequest.STATUS_IN_REVIEW]:
            quote_request.status = QuoteRequest.STATUS_ASSIGNED
        quote_request.save(update_fields=['assigned_to', 'assigned_at', 'status', 'updated_at'])

        AuditLog.log(
            entity=quote_request,
            action=AuditLog.ACTION_UPDATED,
            actor=request.user,
            after_state={'assigned_to': str(assigned_to.id), 'status': quote_request.status},
            request=request
        )

        # Notify the newly assigned seller
        try:
            Notification.notify(
                recipient=assigned_to,
                notification_type=Notification.TYPE_QUOTE_REQUEST,
                title=f'Solicitud asignada: #{quote_request.request_number}',
                message=f'Se te asignó la solicitud de {quote_request.customer_name}',
                entity_type='QuoteRequest',
                entity_id=quote_request.id,
                action_url=f'/dashboard/solicitudes/{quote_request.id}',
            )
        except Exception:
            pass

        return Response(QuoteRequestAdminSerializer(quote_request).data)

    def destroy(self, request, pk=None):
        """Delete quote request (admin only, non-quoted requests only)."""
        if not _is_internal_user(request.user):
            return Response(status=status.HTTP_403_FORBIDDEN)

        # Only admins can delete
        if not (hasattr(request.user, 'role') and request.user.role and request.user.role.name == 'admin'):
            return Response(
                {'error': _('Only administrators can delete requests.')},
                status=status.HTTP_403_FORBIDDEN
            )

        quote_request = self.get_object()

        # Don't allow deleting requests that already have quotes
        deletable_statuses = [
            QuoteRequest.STATUS_PENDING,
            QuoteRequest.STATUS_ASSIGNED,
            QuoteRequest.STATUS_IN_REVIEW,
            QuoteRequest.STATUS_REJECTED,
            QuoteRequest.STATUS_CANCELLED,
        ]
        if quote_request.status not in deletable_statuses:
            return Response(
                {'error': _('Cannot delete a request that has been quoted or accepted.')},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Log deletion
        AuditLog.log(
            entity=quote_request,
            action=AuditLog.ACTION_DELETED,
            actor=request.user,
            before_state=QuoteRequestAdminSerializer(quote_request).data,
            request=request
        )

        quote_request.delete()  # Soft delete via SoftDeleteModel
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['post'])
    def mark_in_review(self, request, pk=None):
        """Mark request as in review (admin/sales)."""
        if not _is_internal_user(request.user):
            return Response(status=status.HTTP_403_FORBIDDEN)

        quote_request = self.get_object()
        quote_request.status = QuoteRequest.STATUS_IN_REVIEW
        quote_request.save(update_fields=['status', 'updated_at'])

        AuditLog.log(
            entity=quote_request,
            action=AuditLog.ACTION_STATE_CHANGED,
            actor=request.user,
            after_state={'status': quote_request.status},
            request=request
        )

        return Response(QuoteRequestAdminSerializer(quote_request).data)

    @action(detail=True, methods=['post'])
    def unmark_in_review(self, request, pk=None):
        """Revert request from in_review back to assigned/pending."""
        if not _is_internal_user(request.user):
            return Response(status=status.HTTP_403_FORBIDDEN)

        quote_request = self.get_object()
        if quote_request.status != QuoteRequest.STATUS_IN_REVIEW:
            return Response(
                {'error': _('Only requests in review can be reverted.')},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Revert to assigned (if has an assignee) or pending
        if quote_request.assigned_to:
            quote_request.status = QuoteRequest.STATUS_ASSIGNED
        else:
            quote_request.status = QuoteRequest.STATUS_PENDING
        quote_request.save(update_fields=['status', 'updated_at'])

        AuditLog.log(
            entity=quote_request,
            action=AuditLog.ACTION_STATE_CHANGED,
            actor=request.user,
            after_state={'status': quote_request.status},
            request=request
        )

        return Response(QuoteRequestAdminSerializer(quote_request).data)

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """Cancel a quote request (admin)."""
        if not _is_internal_user(request.user):
            return Response(status=status.HTTP_403_FORBIDDEN)

        quote_request = self.get_object()
        if quote_request.status in [QuoteRequest.STATUS_ACCEPTED, QuoteRequest.STATUS_CANCELLED]:
            return Response(
                {'error': _('This request cannot be cancelled.')},
                status=status.HTTP_400_BAD_REQUEST
            )

        reason = request.data.get('reason', '')
        quote_request.status = QuoteRequest.STATUS_CANCELLED
        quote_request.save(update_fields=['status', 'updated_at'])

        AuditLog.log(
            entity=quote_request,
            action=AuditLog.ACTION_STATE_CHANGED,
            actor=request.user,
            after_state={'status': quote_request.status, 'reason': reason},
            request=request,
            metadata={'cancelled_reason': reason}
        )

        return Response(QuoteRequestAdminSerializer(quote_request).data)


class QuoteViewSet(viewsets.ModelViewSet):
    """
    ViewSet for quote management.

    Customer endpoints:
        GET /api/v1/quotes/ - List user's quotes
        GET /api/v1/quotes/{id}/ - Quote details
        POST /api/v1/quotes/{id}/accept/ - Accept quote

    Admin endpoints:
        GET /api/v1/admin/quotes/ - List all quotes
        POST /api/v1/admin/quotes/ - Create quote
        PUT /api/v1/admin/quotes/{id}/ - Update quote
        POST /api/v1/admin/quotes/{id}/send/ - Send to customer
    """

    permission_classes = [permissions.IsAuthenticated]
    pagination_class = StandardPagination
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['status']
    search_fields = ['quote_number', 'quote_request__customer_name', 'quote_request__customer_email', 'quote_request__customer_company']

    def get_queryset(self):
        """Return quotes based on user role.

        Visibility rules:
            - Admin: sees ALL quotes.
            - Sales: sees only quotes they created.
            - Customer: sees quotes sent to them.
        """
        user = self.request.user
        if _is_internal_user(user):
            base_qs = Quote.objects.all().select_related(
                'quote_request', 'created_by'
            ).prefetch_related('lines')
            # Sales reps: only quotes they created
            if hasattr(user, 'role') and user.role and user.role.name == 'sales':
                return base_qs.filter(
                    models.Q(created_by=user) |
                    models.Q(quote_request__assigned_to=user)
                )
            return base_qs
        return Quote.objects.filter(
            quote_request__customer_email=self.request.user.email,
            is_deleted=False,
            status__in=[
                Quote.STATUS_SENT,
                Quote.STATUS_VIEWED,
                Quote.STATUS_ACCEPTED,
                Quote.STATUS_REJECTED,
                Quote.STATUS_CHANGES_REQUESTED,
            ]
        ).select_related('quote_request').prefetch_related('lines')

    def get_serializer_class(self):
        """Return appropriate serializer."""
        if self.action == 'create':
            return QuoteCreateSerializer
        if _is_internal_user(self.request.user):
            return QuoteAdminSerializer
        return QuoteSerializer

    def create(self, request, *args, **kwargs):
        """Create quote and return with full serializer.
        
        Sales reps can only create quotes for requests assigned to them,
        unless the request has urgency='high'. Admins can create for any request.
        """
        if not _is_internal_user(request.user):
            return Response(status=status.HTTP_403_FORBIDDEN)

        # Check assignment restriction for sales users (non-admin)
        quote_request_id = request.data.get('quote_request_id')
        if quote_request_id and hasattr(request.user, 'role') and request.user.role and request.user.role.name == 'sales':
            try:
                quote_request = QuoteRequest.objects.get(id=quote_request_id)
                is_assigned_to_me = quote_request.assigned_to_id == request.user.id
                is_urgent = quote_request.urgency == QuoteRequest.URGENCY_HIGH

                if not is_assigned_to_me and not is_urgent:
                    return Response(
                        {'error': _('You can only create quotes for requests assigned to you, unless the request is marked as urgent.')},
                        status=status.HTTP_403_FORBIDDEN
                    )
            except QuoteRequest.DoesNotExist:
                pass

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        quote = serializer.save()

        AuditLog.log(
            entity=quote,
            action=AuditLog.ACTION_CREATED,
            actor=request.user,
            after_state=QuoteAdminSerializer(quote).data,
            request=request
        )

        # Return the quote with QuoteAdminSerializer to include id and other fields
        output_serializer = QuoteAdminSerializer(quote)
        return Response(output_serializer.data, status=status.HTTP_201_CREATED)

    def destroy(self, request, pk=None):
        """Delete quote (admin only, drafts only)."""
        if not _is_internal_user(request.user):
            return Response(status=status.HTTP_403_FORBIDDEN)

        quote = self.get_object()

        # Only allow deleting drafts
        if quote.status != Quote.STATUS_DRAFT:
            return Response(
                {'error': _('Only draft quotes can be deleted.')},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Log deletion
        AuditLog.log(
            entity=quote,
            action=AuditLog.ACTION_DELETED,
            actor=request.user,
            before_state=QuoteAdminSerializer(quote).data,
            request=request
        )

        quote.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    def partial_update(self, request, pk=None):
        """
        PATCH a quote (admin only, drafts only).

        Handles writable nested 'lines' — the frontend sends the full set of
        line items and we replace them atomically.
        """
        if not _is_internal_user(request.user):
            return Response(status=status.HTTP_403_FORBIDDEN)

        quote = self.get_object()

        if quote.status != Quote.STATUS_DRAFT:
            return Response(
                {'error': _('Only draft quotes can be edited.')},
                status=status.HTTP_400_BAD_REQUEST
            )

        before_state = QuoteAdminSerializer(quote).data

        # Separate lines from the rest of the data
        lines_data = request.data.pop('lines', None) if isinstance(request.data, dict) else None

        with transaction.atomic():
            # Update scalar fields via the default serializer
            serializer = self.get_serializer(quote, data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            serializer.save()

            # Replace lines if provided
            if lines_data is not None and isinstance(lines_data, list):
                if not lines_data:
                    return Response(
                        {'error': _('At least one line item is required.')},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Delete existing lines and re-create
                quote.lines.all().delete()

                subtotal = Decimal('0.00')
                for position, line in enumerate(lines_data):
                    quantity = Decimal(str(line.get('quantity', 1)))
                    unit_price = Decimal(str(line.get('unit_price', 0)))
                    line_total = quantity * unit_price

                    QuoteLine.objects.create(
                        quote=quote,
                        concept=line.get('concept', ''),
                        concept_en=line.get('concept_en', ''),
                        description=line.get('description', ''),
                        description_en=line.get('description_en', ''),
                        quantity=quantity,
                        unit=line.get('unit', 'pz'),
                        unit_price=unit_price,
                        line_total=line_total,
                        position=line.get('position', position),
                        service_details=line.get('service_details'),
                    )
                    subtotal += line_total

                # Recalculate totals
                quote.subtotal = subtotal
                quote.tax_amount = subtotal * quote.tax_rate
                quote.total = subtotal + quote.tax_amount
                quote.save(update_fields=['subtotal', 'tax_amount', 'total', 'updated_at'])

            AuditLog.log(
                entity=quote,
                action=AuditLog.ACTION_UPDATED,
                actor=request.user,
                before_state=before_state,
                after_state=QuoteAdminSerializer(quote).data,
                request=request,
            )

        quote.refresh_from_db()
        return Response(QuoteAdminSerializer(quote).data)

    @action(detail=True, methods=['post'])
    def send(self, request, pk=None):
        """Send quote to customer (admin)."""
        if not _is_internal_user(request.user):
            return Response(status=status.HTTP_403_FORBIDDEN)

        quote = self.get_object()

        if quote.status not in [Quote.STATUS_DRAFT]:
            return Response(
                {'error': _('Only draft quotes can be sent.')},
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = QuoteSendSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        with transaction.atomic():
            quote.status = Quote.STATUS_SENT
            quote.sent_at = timezone.now()

            # Update validity if provided
            if 'valid_until' in serializer.validated_data:
                quote.valid_until = serializer.validated_data['valid_until']

            quote.save(update_fields=['status', 'sent_at', 'valid_until', 'updated_at'])

            # Track this send in QuoteResponse so every send appears in the
            # timeline history — especially important when a quote is re-sent
            # after a change request cycle (v1 send, v2 send, etc.).
            version_label = f'v{quote.version}' if quote.version > 1 else ''
            QuoteResponse.objects.create(
                quote=quote,
                action=QuoteResponse.ACTION_SEND,
                comment=version_label,
                responded_by=request.user,
            )

            # Update quote request status
            if quote.quote_request:
                quote.quote_request.status = QuoteRequest.STATUS_QUOTED
                quote.quote_request.save(update_fields=['status', 'updated_at'])

            AuditLog.log(
                entity=quote,
                action=AuditLog.ACTION_STATE_CHANGED,
                actor=request.user,
                after_state={'status': quote.status, 'sent_at': str(quote.sent_at)},
                request=request
            )

        # Regenerate PDF to ensure it reflects current quote data
        from apps.quotes.tasks import generate_quote_pdf
        try:
            generate_quote_pdf(str(quote.id))
            # Also generate English PDF if quote language is English
            if quote.language == 'en':
                generate_quote_pdf(str(quote.id), language='en')
            quote.refresh_from_db()
        except Exception as e:
            logger.warning(f"PDF regeneration failed for {quote.quote_number}: {e}")

        # Send email synchronously so we know if it actually succeeded
        email_sent = False
        email_error = ''
        try:
            send_quote_email_sync(str(quote.id))
            email_sent = True
            logger.info(f"Quote email sent for {quote.quote_number}")
        except Exception as e:
            email_error = str(e)
            logger.error(f"Quote email failed for {quote.quote_number}: {e}")

        response_data = QuoteAdminSerializer(quote).data
        response_data['email_sent'] = email_sent
        if email_error:
            response_data['email_error'] = email_error

        return Response(response_data)

    @action(detail=True, methods=['post'], url_path='resend-email')
    def resend_email(self, request, pk=None):
        """Resend the quote email to the customer."""
        if not _is_internal_user(request.user):
            return Response(status=status.HTTP_403_FORBIDDEN)

        quote = self.get_object()

        if quote.status not in [Quote.STATUS_SENT, Quote.STATUS_VIEWED, Quote.STATUS_CHANGES_REQUESTED]:
            return Response(
                {'error': _('Quote must be sent to resend email.')},
                status=status.HTTP_400_BAD_REQUEST
            )

        email_sent = False
        email_error = ''
        try:
            send_quote_email_sync(str(quote.id))
            email_sent = True
            logger.info(f"Quote email resent for {quote.quote_number}")
        except Exception as e:
            email_error = str(e)
            logger.error(f"Quote email resend failed for {quote.quote_number}: {e}")

        data = {'email_sent': email_sent}
        if email_error:
            data['email_error'] = email_error

        return Response(data, status=status.HTTP_200_OK if email_sent else status.HTTP_502_BAD_GATEWAY)

    @action(detail=True, methods=['post'])
    def accept(self, request, pk=None):
        """Accept a quote (customer)."""
        quote = self.get_object()

        # Security: Verify the authenticated user's email matches the quote's customer email
        if not _is_internal_user(request.user):
            if request.user.email.lower() != quote.customer_email.lower():
                return Response(
                    {'error': _('You are not authorized to accept this quote. Please login with the correct account.')},
                    status=status.HTTP_403_FORBIDDEN
                )

        if quote.status not in [Quote.STATUS_SENT, Quote.STATUS_VIEWED]:
            return Response(
                {'error': _('This quote cannot be accepted.')},
                status=status.HTTP_400_BAD_REQUEST
            )

        if quote.is_expired:
            return Response(
                {'error': _('This quote has expired.')},
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = QuoteAcceptSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        with transaction.atomic():
            quote.status = Quote.STATUS_ACCEPTED
            quote.accepted_at = timezone.now()
            quote.customer_notes = serializer.validated_data.get('notes', '')

            # Handle electronic signature
            signature_data = request.data.get('signature')
            signature_name = request.data.get('signature_name', '')
            if signature_data:
                try:
                    import base64
                    from django.core.files.base import ContentFile
                    # Expect base64 data URI: data:image/png;base64,...
                    if ',' in signature_data:
                        signature_data = signature_data.split(',')[1]
                    img_data = base64.b64decode(signature_data)
                    filename = f"signature_{quote.quote_number}_{timezone.now().strftime('%Y%m%d%H%M%S')}.png"
                    quote.signature_image.save(filename, ContentFile(img_data), save=False)
                    quote.signature_name = signature_name
                    quote.signed_at = timezone.now()
                except Exception as sig_error:
                    logger.warning(f"Could not save signature for {quote.quote_number}: {sig_error}")

            quote.save(update_fields=[
                'status', 'accepted_at', 'customer_notes',
                'signature_image', 'signature_name', 'signed_at', 'updated_at'
            ])

            # Update quote request status
            if quote.quote_request:
                quote.quote_request.status = QuoteRequest.STATUS_ACCEPTED
                quote.quote_request.save(update_fields=['status', 'updated_at'])

            AuditLog.log(
                entity=quote,
                action=AuditLog.ACTION_STATE_CHANGED,
                actor=request.user,
                after_state={'status': quote.status, 'accepted_at': str(quote.accepted_at)},
                request=request
            )

            # Notify sales team about accepted quote
            try:
                send_quote_accepted_notification.delay(str(quote.id))
            except Exception:
                pass  # Don't fail the request if notification fails

            # In-app notification to quote owner + admins
            try:
                Notification.notify_owner_and_admins(
                    owner=quote.created_by,
                    notification_type=Notification.TYPE_QUOTE_ACCEPTED,
                    title=f'Cotización #{quote.quote_number} aceptada',
                    message=f'{quote.customer_name} aceptó la cotización.',
                    entity_type='Quote',
                    entity_id=quote.id,
                    action_url=f'/dashboard/cotizaciones/{quote.id}',
                )
            except Exception:
                pass

        return Response(QuoteSerializer(quote).data)

    @action(detail=True, methods=['post'])
    def convert_to_order(self, request, pk=None):
        """
        Convert an accepted quote into an order.

        Only staff (admin/sales) can convert quotes.
        The quote must be in 'accepted' status.
        """
        quote = self.get_object()

        if not _is_internal_user(request.user):
            return Response(
                {'error': _('Only staff can convert quotes to orders.')},
                status=status.HTTP_403_FORBIDDEN
            )

        if quote.status != Quote.STATUS_ACCEPTED:
            return Response(
                {'error': _('Only accepted quotes can be converted to orders.')},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check if quote has a customer user linked
        if not quote.customer:
            return Response(
                {'error': _('This quote does not have a registered customer linked. Please link a customer before converting to order.')},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check if already converted
        from apps.orders.models import Order, OrderLine, OrderStatusHistory
        if Order.objects.filter(quote=quote).exists():
            return Response(
                {'error': _('This quote has already been converted to an order.')},
                status=status.HTTP_400_BAD_REQUEST
            )

        payment_method = request.data.get('payment_method', 'bank_transfer')
        notes = request.data.get('notes', '')

        with transaction.atomic():
            # Create order from quote data
            order = Order.objects.create(
                user=quote.customer,
                status=Order.STATUS_PENDING_PAYMENT,
                subtotal=quote.subtotal,
                tax_rate=quote.tax_rate,
                tax_amount=quote.tax_amount,
                total=quote.total,
                currency=quote.currency,
                payment_method=payment_method,
                notes=notes,
                internal_notes=_(
                    'Converted from quote %(quote_number)s'
                ) % {'quote_number': quote.quote_number},
                quote=quote,
                # Delivery info inherited from quote
                delivery_method=quote.delivery_method,
                pickup_branch=quote.pickup_branch,
                delivery_address=quote.delivery_address,
            )

            # Create order lines from quote lines
            for line in quote.lines.all():
                OrderLine.objects.create(
                    order=order,
                    sku=f'Q-{quote.quote_number}-{line.position}',
                    name=line.concept,
                    variant_name=line.description[:255] if line.description else '',
                    quantity=int(line.quantity),
                    unit_price=line.unit_price,
                    line_total=line.line_total,
                    metadata={
                        'quote_line_id': str(line.id),
                        'unit': line.unit,
                        'original_quantity': str(line.quantity),
                    }
                )

            # Create initial status history
            OrderStatusHistory.objects.create(
                order=order,
                from_status=Order.STATUS_DRAFT,
                to_status=Order.STATUS_PENDING_PAYMENT,
                changed_by=request.user,
                notes=_(
                    'Order created from quote %(quote_number)s'
                ) % {'quote_number': quote.quote_number}
            )

            # Update quote status to converted
            quote.status = Quote.STATUS_CONVERTED
            quote.save(update_fields=['status', 'updated_at'])

            # Update quote request status if exists
            if quote.quote_request and quote.quote_request.status != QuoteRequest.STATUS_ACCEPTED:
                quote.quote_request.status = QuoteRequest.STATUS_ACCEPTED
                quote.quote_request.save(update_fields=['status', 'updated_at'])

            AuditLog.log(
                entity=quote,
                action=AuditLog.ACTION_STATE_CHANGED,
                actor=request.user,
                before_state={'status': Quote.STATUS_ACCEPTED},
                after_state={
                    'status': Quote.STATUS_CONVERTED,
                    'order_id': str(order.id),
                    'order_number': order.order_number,
                },
                request=request
            )

        # In-app notification: order created from quote
        try:
            Notification.notify_owner_and_admins(
                owner=quote.created_by,
                notification_type=Notification.TYPE_ORDER_CREATED,
                title=f'Pedido #{order.order_number} creado',
                message=f'Convertido desde cotización #{quote.quote_number} — {quote.customer_name}.',
                entity_type='Order',
                entity_id=order.id,
                action_url=f'/dashboard/pedidos/{order.id}',
            )
        except Exception:
            pass

        # Send order confirmation email to customer
        try:
            import threading as _threading
            _threading.Thread(
                target=send_order_confirmation_email,
                args=(str(quote.id), order.order_number),
                daemon=True
            ).start()
        except Exception:
            pass

        from apps.orders.serializers import OrderSerializer as OrderSerializerFull
        return Response({
            'quote': QuoteSerializer(quote).data,
            'order': OrderSerializerFull(order).data,
        }, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Reject a quote (customer)."""
        quote = self.get_object()

        # Security: Verify the authenticated user's email matches the quote's customer email
        if not _is_internal_user(request.user):
            if request.user.email.lower() != quote.customer_email.lower():
                return Response(
                    {'error': _('You are not authorized to reject this quote. Please login with the correct account.')},
                    status=status.HTTP_403_FORBIDDEN
                )

        if quote.status not in [Quote.STATUS_SENT, Quote.STATUS_VIEWED]:
            return Response(
                {'error': _('This quote cannot be rejected.')},
                status=status.HTTP_400_BAD_REQUEST
            )

        reason = request.data.get('reason', '')

        with transaction.atomic():
            quote.status = Quote.STATUS_REJECTED
            quote.customer_notes = reason
            quote.save(update_fields=['status', 'customer_notes', 'updated_at'])

            # Update quote request status
            if quote.quote_request:
                quote.quote_request.status = QuoteRequest.STATUS_REJECTED
                quote.quote_request.save(update_fields=['status', 'updated_at'])

            AuditLog.log(
                entity=quote,
                action=AuditLog.ACTION_STATE_CHANGED,
                actor=request.user,
                after_state={'status': quote.status, 'reason': reason},
                request=request
            )

            # In-app notification: quote rejected
            try:
                Notification.notify_owner_and_admins(
                    owner=quote.created_by,
                    notification_type=Notification.TYPE_QUOTE_REJECTED,
                    title=f'Cotización #{quote.quote_number} rechazada',
                    message=f'{quote.customer_name}: {reason[:120]}' if reason else f'{quote.customer_name} rechazó la cotización.',
                    entity_type='Quote',
                    entity_id=quote.id,
                    action_url=f'/dashboard/cotizaciones/{quote.id}',
                )
            except Exception:
                pass

        return Response(QuoteSerializer(quote).data)

    @action(detail=True, methods=['post'])
    def duplicate(self, request, pk=None):
        """Duplicate a quote (admin)."""
        if not _is_internal_user(request.user):
            return Response(status=status.HTTP_403_FORBIDDEN)

        original = self.get_object()

        with transaction.atomic():
            # Create new quote — copy ALL fields from original
            new_quote = Quote.objects.create(
                quote_request=original.quote_request,
                created_by=request.user,
                status=Quote.STATUS_DRAFT,
                # Customer
                customer=original.customer,
                customer_name=original.customer_name,
                customer_email=original.customer_email,
                customer_company=original.customer_company,
                customer_phone=original.customer_phone,
                # Financials
                subtotal=original.subtotal,
                tax_rate=original.tax_rate,
                tax_amount=original.tax_amount,
                total=original.total,
                currency=original.currency,
                # Payment
                payment_mode=original.payment_mode,
                deposit_percentage=original.deposit_percentage,
                payment_methods=original.payment_methods,
                payment_conditions=original.payment_conditions,
                # Delivery
                delivery_time_text=original.delivery_time_text,
                estimated_delivery_date=original.estimated_delivery_date,
                # Terms & notes
                terms=original.terms,
                terms_en=original.terms_en,
                customer_notes=original.customer_notes,
                included_services=original.included_services,
                internal_notes=f"Duplicado de {original.quote_number}",
                # Language & validity
                language=original.language,
                valid_until=original.valid_until,
            )

            # Copy lines
            for line in original.lines.all():
                QuoteLine.objects.create(
                    quote=new_quote,
                    concept=line.concept,
                    concept_en=line.concept_en,
                    description=line.description,
                    description_en=line.description_en,
                    quantity=line.quantity,
                    unit=line.unit,
                    unit_price=line.unit_price,
                    line_total=line.line_total,
                    position=line.position
                )

            AuditLog.log(
                entity=new_quote,
                action=AuditLog.ACTION_CREATED,
                actor=request.user,
                after_state=QuoteAdminSerializer(new_quote).data,
                request=request,
                metadata={'duplicated_from': str(original.id)}
            )

        return Response(
            QuoteAdminSerializer(new_quote).data,
            status=status.HTTP_201_CREATED
        )

    @action(detail=True, methods=['get'])
    def pdf(self, request, pk=None):
        """Download quote PDF."""
        quote = self.get_object()

        # Generate PDF on the fly if missing
        if not quote.pdf_file:
            try:
                from apps.quotes.tasks import generate_quote_pdf
                # Call directly — no Celery broker needed
                generate_quote_pdf(str(quote.id))
                quote.refresh_from_db()
                logger.info(f'PDF generated on-demand for {quote.quote_number}')
            except Exception as e:
                logger.error(f'Error generating PDF for quote {quote.quote_number}: {type(e).__name__}: {e}', exc_info=True)
                return Response(
                    {'error': f'Could not generate PDF: {type(e).__name__}: {str(e)[:200]}'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        if not quote.pdf_file:
            logger.error(f'PDF still empty after generation for quote {quote.quote_number}')
            return Response(
                {'error': _('PDF not available. Generation may have failed.')},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            pdf_data = quote.pdf_file.read()
            response = HttpResponse(pdf_data, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="cotizacion_{quote.quote_number}.pdf"'
            return response
        except Exception as e:
            logger.error(f'Error reading PDF for quote {quote.quote_number}: {type(e).__name__}: {e}', exc_info=True)
            return Response(
                {'error': f'Error reading PDF file: {type(e).__name__}: {str(e)[:200]}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['patch'], url_path='internal-notes')
    def update_internal_notes(self, request, pk=None):
        """Update internal notes for a quote (admin/sales) regardless of status."""
        if not _is_internal_user(request.user):
            return Response(status=status.HTTP_403_FORBIDDEN)

        quote = self.get_object()
        notes = request.data.get('internal_notes', '')
        quote.internal_notes = notes
        quote.save(update_fields=['internal_notes', 'updated_at'])
        return Response(QuoteAdminSerializer(quote).data)

    @action(detail=True, methods=['post'], url_path='regenerate-pdf')
    def regenerate_pdf(self, request, pk=None):
        """Regenerate quote PDF (admin)."""
        if not _is_internal_user(request.user):
            return Response(status=status.HTTP_403_FORBIDDEN)

        quote = self.get_object()

        # Trigger PDF regeneration (direct call — no broker needed)
        from apps.quotes.tasks import generate_quote_pdf
        generate_quote_pdf(str(quote.id))

        return Response({'message': _('PDF regenerated successfully.')})

    @action(detail=True, methods=['post'], url_path='upload-attachment')
    def upload_attachment(self, request, pk=None):
        """Upload a file attachment to a quote (seller/admin)."""
        if not _is_internal_user(request.user):
            return Response(status=status.HTTP_403_FORBIDDEN)

        quote = self.get_object()
        file = request.FILES.get('file')
        if not file:
            return Response(
                {'error': _('No file provided.')},
                status=status.HTTP_400_BAD_REQUEST
            )

        from .serializers import QuoteAttachmentSerializer
        serializer = QuoteAttachmentSerializer(data={'file': file})
        serializer.is_valid(raise_exception=True)

        attachment = QuoteAttachment.objects.create(
            quote=quote,
            file=file,
            filename=file.name,
            file_type=file.content_type,
            file_size=file.size,
            description=request.data.get('description', ''),
        )

        return Response(
            QuoteAttachmentSerializer(attachment).data,
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=['delete'], url_path='delete-attachment/(?P<attachment_id>[^/.]+)')
    def delete_attachment(self, request, pk=None, attachment_id=None):
        """Delete a file attachment from a quote (seller/admin)."""
        if not _is_internal_user(request.user):
            return Response(status=status.HTTP_403_FORBIDDEN)

        quote = self.get_object()
        try:
            attachment = QuoteAttachment.objects.get(id=attachment_id, quote=quote)
        except QuoteAttachment.DoesNotExist:
            return Response(
                {'error': _('Attachment not found.')},
                status=status.HTTP_404_NOT_FOUND,
            )
        attachment.file.delete(save=False)
        attachment.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """Export filtered quotes to Excel."""
        if not _is_internal_user(request.user):
            return Response(status=status.HTTP_403_FORBIDDEN)

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO

        qs = self.get_queryset()
        # Apply filters
        status_filter = request.query_params.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)
        search = request.query_params.get('search')
        if search:
            qs = qs.filter(
                models.Q(quote_number__icontains=search) |
                models.Q(customer_name__icontains=search) |
                models.Q(customer_email__icontains=search) |
                models.Q(customer_company__icontains=search)
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Cotizaciones'

        # Header style
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0DA3EF', end_color='0DA3EF', fill_type='solid')

        headers = [
            'Número', 'Estado', 'Versión', 'Cliente', 'Email', 'Empresa',
            'Subtotal', 'IVA', 'Total', 'Moneda', 'Creada', 'Enviada',
            'Válida hasta', 'Aceptada', 'Creada por',
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row_idx, quote in enumerate(qs[:1000], 2):
            ws.cell(row=row_idx, column=1, value=quote.quote_number)
            ws.cell(row=row_idx, column=2, value=quote.get_status_display())
            ws.cell(row=row_idx, column=3, value=quote.version)
            ws.cell(row=row_idx, column=4, value=quote.customer_name)
            ws.cell(row=row_idx, column=5, value=quote.customer_email)
            ws.cell(row=row_idx, column=6, value=quote.customer_company or '')
            ws.cell(row=row_idx, column=7, value=float(quote.subtotal))
            ws.cell(row=row_idx, column=8, value=float(quote.tax_amount))
            ws.cell(row=row_idx, column=9, value=float(quote.total))
            ws.cell(row=row_idx, column=10, value=quote.currency)
            ws.cell(row=row_idx, column=11, value=quote.created_at.strftime('%Y-%m-%d %H:%M') if quote.created_at else '')
            ws.cell(row=row_idx, column=12, value=quote.sent_at.strftime('%Y-%m-%d %H:%M') if quote.sent_at else '')
            ws.cell(row=row_idx, column=13, value=quote.valid_until.strftime('%Y-%m-%d') if quote.valid_until else '')
            ws.cell(row=row_idx, column=14, value=quote.accepted_at.strftime('%Y-%m-%d %H:%M') if quote.accepted_at else '')
            ws.cell(row=row_idx, column=15, value=quote.created_by.full_name if quote.created_by else '')

        # Auto-width columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="cotizaciones_export.xlsx"'
        return response


class QuotePublicView(APIView):
    """
    Public view for quotes via secure token.

    GET /api/v1/quotes/view/{token}/
    """

    permission_classes = [permissions.AllowAny]

    def get(self, request, token):
        """View quote via secure token."""
        try:
            quote = Quote.objects.select_related(
                'quote_request', 'quote_request__catalog_item'
            ).prefetch_related(
                'lines', 'attachments', 'quote_request__attachments'
            ).get(
                token=token,
                status__in=[
                    Quote.STATUS_SENT,
                    Quote.STATUS_VIEWED,
                    Quote.STATUS_ACCEPTED,
                    Quote.STATUS_REJECTED,
                    Quote.STATUS_CHANGES_REQUESTED,
                ]
            )
        except Quote.DoesNotExist:
            return Response(
                {'error': _('Quote not found.')},
                status=status.HTTP_404_NOT_FOUND
            )

        # Mark as viewed if first time
        if quote.status == Quote.STATUS_SENT:
            quote.status = Quote.STATUS_VIEWED
            quote.viewed_at = timezone.now()
            quote.save(update_fields=['status', 'viewed_at', 'updated_at'])

            AuditLog.log(
                entity=quote,
                action=AuditLog.ACTION_VIEWED,
                request=request,
                metadata={'viewed_via': 'token_link'}
            )

        return Response(QuoteSerializer(quote).data)


class QuoteChangeRequestView(APIView):
    """
    Public endpoint for customers to request changes to a quote.

    GET /api/v1/quotes/view/{token}/change-request/
        - Returns pending change requests for this quote

    POST /api/v1/quotes/view/{token}/change-request/
        - Creates a structured change request with proposed line modifications

    This does not require authentication. It creates a change request
    record that the sales team can review.
    """

    permission_classes = [permissions.AllowAny]

    def get(self, request, token):
        """Get existing change requests for this quote."""
        try:
            quote = Quote.objects.get(
                token=token,
                status__in=[
                    Quote.STATUS_SENT,
                    Quote.STATUS_VIEWED,
                    Quote.STATUS_CHANGES_REQUESTED,
                ]
            )
        except Quote.DoesNotExist:
            return Response(
                {'error': _('Quote not found.')},
                status=status.HTTP_404_NOT_FOUND
            )

        # Get all change requests for this quote
        change_requests = quote.change_requests.all().order_by('-created_at')
        serializer = QuoteChangeRequestSerializer(change_requests, many=True)

        return Response({
            'quote_number': quote.quote_number,
            'quote_status': quote.status,
            'change_requests': serializer.data,
        })

    def post(self, request, token):
        """Submit a structured change request for a quote."""
        try:
            quote = Quote.objects.select_related(
                'quote_request', 'created_by'
            ).prefetch_related('lines').get(
                token=token,
                status__in=[
                    Quote.STATUS_SENT,
                    Quote.STATUS_VIEWED,
                ]
            )
        except Quote.DoesNotExist:
            return Response(
                {'error': _('Quote not found or cannot be modified.')},
                status=status.HTTP_404_NOT_FOUND
            )

        if quote.is_expired:
            return Response(
                {'error': _('This quote has expired.')},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check if there's already a pending change request
        pending_request = quote.change_requests.filter(
            status=QuoteChangeRequest.STATUS_PENDING
        ).first()
        if pending_request:
            return Response(
                {'error': _('There is already a pending change request for this quote. Please wait for it to be reviewed.')},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validate and create the change request
        # Handle multipart/form-data: proposed_lines arrives as JSON string
        import json as _json
        raw_data = request.data

        # Build a plain dict for the serializer
        data = {}
        proposed_raw = raw_data.get('proposed_lines')
        if isinstance(proposed_raw, str):
            try:
                data['proposed_lines'] = _json.loads(proposed_raw)
            except (ValueError, TypeError):
                return Response(
                    {'error': _('Invalid proposed_lines format.')},
                    status=status.HTTP_400_BAD_REQUEST
                )
        elif isinstance(proposed_raw, list):
            data['proposed_lines'] = proposed_raw
        else:
            return Response(
                {'error': _('proposed_lines is required.')},
                status=status.HTTP_400_BAD_REQUEST
            )

        if raw_data.get('customer_comments'):
            data['customer_comments'] = raw_data['customer_comments']

        # Collect uploaded files
        attachments = request.FILES.getlist('attachments')
        if attachments:
            data['attachments'] = attachments

        serializer = QuoteChangeRequestCreateSerializer(
            data=data,
            context={'quote': quote, 'request': request}
        )
        serializer.is_valid(raise_exception=True)

        # Validate delete actions - ensure not all lines are being deleted
        proposed_lines = serializer.validated_data['proposed_lines']
        delete_ids = {
            str(line['id']) for line in proposed_lines
            if line.get('action') == 'delete' and line.get('id')
        }
        add_lines = [line for line in proposed_lines if line.get('action') == 'add']
        existing_line_ids = {str(line.id) for line in quote.lines.all()}

        # Check if all existing lines are being deleted and no new lines are added
        remaining_lines = existing_line_ids - delete_ids
        if not remaining_lines and not add_lines:
            return Response(
                {'error': _('You cannot delete all items from the quote. Please add at least one item or modify existing ones.')},
                status=status.HTTP_400_BAD_REQUEST
            )

        with transaction.atomic():
            change_request = serializer.save()

            # Update quote status
            quote.status = Quote.STATUS_CHANGES_REQUESTED
            quote.save(update_fields=['status', 'updated_at'])

            # Log the change request
            AuditLog.log(
                entity=quote,
                action=AuditLog.ACTION_UPDATED,
                request=request,
                metadata={
                    'type': 'change_request',
                    'change_request_id': str(change_request.id),
                    'customer_email': quote.customer_email,
                    'customer_name': quote.customer_name,
                    'changes_summary': change_request.get_changes_summary(),
                }
            )

            # In-app notification: change request
            try:
                changes = change_request.get_changes_summary()
                parts = []
                if changes.get('added'):
                    parts.append(f"{changes['added']} nuevas")
                if changes.get('modified'):
                    parts.append(f"{changes['modified']} modificadas")
                if changes.get('deleted'):
                    parts.append(f"{changes['deleted']} eliminadas")
                changes_text = ', '.join(parts) if parts else 'ver detalle'

                Notification.notify_owner_and_admins(
                    owner=quote.created_by,
                    notification_type=Notification.TYPE_CHANGE_REQUEST,
                    title=f'Solicitud de cambios #{quote.quote_number}',
                    message=f'{quote.customer_name} — {changes_text}',
                    entity_type='QuoteChangeRequest',
                    entity_id=change_request.id,
                    action_url=f'/dashboard/cotizaciones/{quote.id}/cambios/{change_request.id}',
                )
            except Exception:
                pass

        # Send notification email to sales team
        from django.core.mail import send_mail
        from django.conf import settings

        try:
            sales_email = quote.created_by.email if quote.created_by else settings.DEFAULT_FROM_EMAIL
            changes = change_request.get_changes_summary()
            changes_text = []
            if changes.get('added'):
                changes_text.append(f"- {changes['added']} elemento(s) agregado(s)")
            if changes.get('modified'):
                changes_text.append(f"- {changes['modified']} elemento(s) modificado(s)")
            if changes.get('deleted'):
                changes_text.append(f"- {changes['deleted']} elemento(s) eliminado(s)")

            send_mail(
                subject=f'Solicitud de cambios en cotización #{quote.quote_number}',
                message=f'''
El cliente ha solicitado cambios en la cotización #{quote.quote_number}:

Cliente: {quote.customer_name}
Email: {quote.customer_email}
Empresa: {quote.customer_company or 'N/A'}

Cambios solicitados:
{chr(10).join(changes_text) or "Ver detalles en el sistema"}

Comentarios del cliente:
{change_request.customer_comments or "Sin comentarios adicionales"}

---
Revisar solicitud: {settings.FRONTEND_URL}/es/ventas/cotizaciones/{quote.id}/cambios/{change_request.id}
''',
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[sales_email],
                fail_silently=True,
            )
        except Exception:
            pass  # Don't fail the request if email fails

        return Response({
            'message': _('Your change request has been submitted. We will review it and contact you soon.'),
            'change_request': QuoteChangeRequestSerializer(change_request).data,
        }, status=status.HTTP_201_CREATED)


class QuotePublicPdfView(APIView):
    """
    Public endpoint to download quote PDF using token.

    GET /api/v1/quotes/view/{token}/pdf/
    """

    permission_classes = [permissions.AllowAny]

    def get(self, request, token):
        """Download quote PDF via secure token."""
        try:
            quote = Quote.objects.get(
                token=token,
                status__in=[
                    Quote.STATUS_SENT,
                    Quote.STATUS_VIEWED,
                    Quote.STATUS_ACCEPTED,
                    Quote.STATUS_REJECTED,
                    Quote.STATUS_CHANGES_REQUESTED,
                ]
            )
        except Quote.DoesNotExist:
            return Response(
                {'error': _('Quote not found.')},
                status=status.HTTP_404_NOT_FOUND
            )

        # Generate PDF on the fly if missing
        if not quote.pdf_file:
            try:
                from apps.quotes.tasks import generate_quote_pdf
                generate_quote_pdf(str(quote.id))
                quote.refresh_from_db()
                logger.info(f'Public PDF generated on-demand for {quote.quote_number}')
            except Exception as e:
                logger.error(f'Error generating PDF for quote {quote.quote_number}: {type(e).__name__}: {e}', exc_info=True)
                return Response(
                    {'error': f'Could not generate PDF: {type(e).__name__}: {str(e)[:200]}'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        if not quote.pdf_file:
            logger.error(f'Public PDF still empty after generation for quote {quote.quote_number}')
            return Response(
                {'error': _('PDF not available. Generation may have failed.')},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            pdf_data = quote.pdf_file.read()
            response = HttpResponse(pdf_data, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="cotizacion_{quote.quote_number}.pdf"'
            return response
        except Exception as e:
            logger.error(f'Error reading PDF for quote {quote.quote_number}: {type(e).__name__}: {e}', exc_info=True)
            return Response(
                {'error': f'Error reading PDF file: {type(e).__name__}: {str(e)[:200]}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class QuotePublicRejectView(APIView):
    """
    Public endpoint for customers to reject a quote without authentication.

    POST /api/v1/quotes/view/{token}/reject/
    """

    permission_classes = [permissions.AllowAny]

    def post(self, request, token):
        """Reject a quote via secure token (no auth required)."""
        try:
            quote = Quote.objects.select_related('quote_request').get(
                token=token,
                status__in=[
                    Quote.STATUS_SENT,
                    Quote.STATUS_VIEWED,
                ]
            )
        except Quote.DoesNotExist:
            return Response(
                {'error': _('Quote not found or cannot be rejected.')},
                status=status.HTTP_404_NOT_FOUND
            )

        if quote.is_expired:
            return Response(
                {'error': _('This quote has expired.')},
                status=status.HTTP_400_BAD_REQUEST
            )

        reason = request.data.get('reason', '').strip()
        if not reason:
            return Response(
                {'error': _('Please provide a reason for rejection.')},
                status=status.HTTP_400_BAD_REQUEST
            )

        with transaction.atomic():
            quote.status = Quote.STATUS_REJECTED
            quote.customer_notes = reason
            quote.save(update_fields=['status', 'customer_notes', 'updated_at'])

            # Update quote request status
            if quote.quote_request:
                quote.quote_request.status = QuoteRequest.STATUS_REJECTED
                quote.quote_request.save(update_fields=['status', 'updated_at'])

            AuditLog.log(
                entity=quote,
                action=AuditLog.ACTION_STATE_CHANGED,
                request=request,
                metadata={
                    'status': quote.status,
                    'reason': reason,
                    'rejected_via': 'public_token'
                }
            )

            # In-app notification: quote rejected (public)
            try:
                Notification.notify_owner_and_admins(
                    owner=quote.created_by,
                    notification_type=Notification.TYPE_QUOTE_REJECTED,
                    title=f'Cotización #{quote.quote_number} rechazada',
                    message=f'{quote.customer_name}: {reason[:120]}' if reason else f'{quote.customer_name} rechazó la cotización.',
                    entity_type='Quote',
                    entity_id=quote.id,
                    action_url=f'/dashboard/cotizaciones/{quote.id}',
                )
            except Exception:
                pass

        # Notify sales team
        from django.core.mail import send_mail
        from django.conf import settings

        try:
            sales_email = quote.created_by.email if quote.created_by else settings.DEFAULT_FROM_EMAIL
            send_mail(
                subject=f'Cotización #{quote.quote_number} rechazada',
                message=f'''
El cliente ha rechazado la cotización #{quote.quote_number}:

Cliente: {quote.customer_name}
Email: {quote.customer_email}
Empresa: {quote.customer_company or 'N/A'}

Motivo del rechazo:
{reason}

---
Ver cotización: {settings.FRONTEND_URL}/es/ventas/cotizaciones/{quote.id}
''',
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[sales_email],
                fail_silently=True,
            )
        except Exception:
            pass

        return Response({
            'message': _('Quote rejected. Thank you for your response.'),
            'status': quote.status,
        })


class QuoteChangeRequestViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for managing quote change requests (admin/sales).

    GET /api/v1/quotes/change-requests/ - List all change requests
    GET /api/v1/quotes/change-requests/{id}/ - Change request details
    POST /api/v1/quotes/change-requests/{id}/review/ - Approve or reject
    """

    permission_classes = [permissions.IsAuthenticated]
    pagination_class = StandardPagination
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['status', 'quote']
    search_fields = ['customer_name', 'customer_email', 'customer_comments', 'quote__quote_number']
    serializer_class = QuoteChangeRequestSerializer

    def get_queryset(self):
        """Return change requests based on user role."""
        user = self.request.user
        base_qs = QuoteChangeRequest.objects.select_related(
            'quote', 'quote__created_by', 'reviewed_by'
        ).order_by('-created_at')

        if _is_internal_user(user):
            # Admin sees all change requests
            if hasattr(user, 'role') and user.role and user.role.name == 'admin':
                return base_qs

            # Sales reps only see change requests for quotes they created
            return base_qs.filter(quote__created_by=user)

        # Regular users don't have access to this viewset
        return QuoteChangeRequest.objects.none()

    @action(detail=True, methods=['post'])
    def review(self, request, pk=None):
        """Review a change request (approve or reject)."""
        if not _is_internal_user(request.user):
            return Response(status=status.HTTP_403_FORBIDDEN)

        change_request = self.get_object()

        # Sales reps can only review change requests for quotes they created
        user = request.user
        is_admin = hasattr(user, 'role') and user.role and user.role.name == 'admin'
        if not is_admin:
            if change_request.quote.created_by != user:
                return Response(
                    {'error': _('You can only review change requests for quotes you created.')},
                    status=status.HTTP_403_FORBIDDEN
                )

        if change_request.status != QuoteChangeRequest.STATUS_PENDING:
            return Response(
                {'error': _('This change request has already been reviewed.')},
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = QuoteChangeRequestReviewSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        action_type = serializer.validated_data['action']
        notes = serializer.validated_data.get('notes', '')

        with transaction.atomic():
            if action_type == 'approve':
                change_request.approve(request.user, notes)
                message = _('Change request approved. The quote has been returned to draft status for editing.')
            else:
                change_request.reject(request.user, notes)
                message = _('Change request rejected. The quote has been returned to viewed status.')

            AuditLog.log(
                entity=change_request.quote,
                action=AuditLog.ACTION_UPDATED,
                actor=request.user,
                after_state={
                    'change_request_id': str(change_request.id),
                    'review_action': action_type,
                    'review_notes': notes,
                },
                request=request,
                metadata={'type': 'change_request_review'}
            )

        # Send notification to customer
        from django.core.mail import send_mail
        from django.conf import settings

        try:
            quote = change_request.quote
            if action_type == 'approve':
                subject = f'Solicitud de cambios aprobada - Cotización #{quote.quote_number}'
                body = f'''
Estimado/a {quote.customer_name},

Su solicitud de cambios para la cotización #{quote.quote_number} ha sido aprobada.

Nuestro equipo está trabajando en actualizar su cotización y le enviaremos una nueva versión pronto.

{f"Notas del equipo: {notes}" if notes else ""}

Gracias por su preferencia.

Atentamente,
Agencia MCD
'''
            else:
                subject = f'Solicitud de cambios no aprobada - Cotización #{quote.quote_number}'
                body = f'''
Estimado/a {quote.customer_name},

Lamentamos informarle que su solicitud de cambios para la cotización #{quote.quote_number} no pudo ser aprobada.

{f"Motivo: {notes}" if notes else ""}

Si tiene alguna pregunta, no dude en contactarnos.

Atentamente,
Agencia MCD
'''

            send_mail(
                subject=subject,
                message=body,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[quote.customer_email],
                fail_silently=True,
            )
        except Exception:
            pass

        return Response({
            'message': message,
            'change_request': QuoteChangeRequestSerializer(change_request).data,
        })

    @action(detail=False, methods=['get'])
    def pending(self, request):
        """Get all pending change requests."""
        if not _is_internal_user(request.user):
            return Response(status=status.HTTP_403_FORBIDDEN)

        pending = self.get_queryset().filter(status=QuoteChangeRequest.STATUS_PENDING)
        page = self.paginate_queryset(pending)

        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(pending, many=True)
        return Response(serializer.data)


class QuoteCronView(APIView):
    """
    Protected endpoint for external cron scheduler.
    Runs expired-quotes check + reminder emails.
    Authenticate with ?key=<CRON_SECRET_KEY> or Authorization header.
    """
    permission_classes = [permissions.AllowAny]

    def _check_cron_key(self, request):
        from django.conf import settings
        expected_key = getattr(settings, 'CRON_SECRET_KEY', None)
        if not expected_key:
            return False
        key = request.query_params.get('key') or request.headers.get('X-Cron-Key', '')
        return key == expected_key

    def post(self, request):
        if not self._check_cron_key(request):
            return Response({'error': 'Invalid key'}, status=status.HTTP_403_FORBIDDEN)

        from .tasks import (
            expire_old_quotes, send_quote_reminders,
            check_expiring_quotes, check_unattended_requests,
        )

        expired_count = expire_old_quotes()
        reminder_count = send_quote_reminders()
        expiring_notifs = check_expiring_quotes()
        unattended_notifs = check_unattended_requests()

        return Response({
            'expired': expired_count,
            'reminders': reminder_count,
            'expiring_notifications': expiring_notifs,
            'unattended_notifications': unattended_notifs,
            'timestamp': timezone.now().isoformat(),
        })
